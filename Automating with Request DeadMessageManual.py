from requests import post, exceptions
from datetime import datetime, timedelta
import xmltodict, csv, json, time, os, sys, ctypes, ntpath, art
from openpyxl import Workbook, load_workbook
from shutil import copyfile
import threading


class Automate:
    def __init__(self):
        art.tprint("BKS Solutions")
        self.book = Workbook()
        self.sheet = self.book.active
        self.sheet.title = "Dead Messages"
        self.sheet.append(["TASKID", "ERRORMSG", "CREATEUSER", "ENTITYINFO", "TERM", "DISCARDED/RECOVERED", "REASON"])
        self.sheet2 = self.book.create_sheet("REFRESH")
        self.sheet2.append(['TERM', "TOTALBEFORE", "TOTALAFTER"])
        self.headers = {'Connection': 'keep-alive', 'Sec-Fetch-Dest': 'empty',
                        'User-Agent': 'Mozilla/5.0 (Windows NT 6.2; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/80.0.3987.149 Safari/537.36',
                        'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8', 'Accept': '*/*',
                        'Origin': 'https://mlog.unitymedia.de', 'Sec-Fetch-Site': 'same-origin',
                        'Sec-Fetch-Mode': 'cors', 'Referer': 'https://mlog.unitymedia.de/mRic-app/mric',
                        'Accept-Language': 'en-US,en;q=0.9'}
        if not os.path.exists('files'):
            os.mkdir('files')
        self.filename = f'files/DeadMessages_{datetime.now().strftime("%Y %B %d %H-%M")}.xlsx'

    def login(self, creds):
        self.user = creds.get('username')
        self.passw = creds.get('password')
        r = post("https://mlog.unitymedia.de/getprofiles/ajax-getprofiles",
                 data={"username": self.user, "password": self.passw, "client": "mRIC"}, headers=self.headers)
        text = xmltodict.parse(r.text)
        self.session = text['profiles']['session']
        if text['profiles']['profile'].__class__ == list:
            self.profile = text['profiles']['profile'][-1]
        else:
            self.profile = text['profiles']['profile']
        return self.session, self.profile

    def request(self, query, params):
        self.headers['Content-Type'] = 'text/xml;charset="UTF-8"'
        data = f"""<?xml version="1.0"?><QueryRequest><query>{query}</query><username>{self.user}</username><token>{self.session}</token><profile>{self.profile}</profile><resultJSON>true</resultJSON><blobToFile>false</blobToFile><retrieveCount>true</retrieveCount>{params}</QueryRequest>"""
        allFields = []
        dates = []
        headers = {}
        with post(f'https://mlog.unitymedia.de/queryExecution/queryExecution;{query}', headers=self.headers,
                  data=data) as r:
            for head in r.json()['metadata']:
                headers[head['name']] = head['colid']
                if head['type'] == 'dateTime':
                    dates.append(head['colid'])
            for row in r.json()['data']:
                fields = {str(i): '' for i in range(1, int(r.json()['metadata'][-1]['colid']) + 1)}
                for cell in row:
                    if cell['colid'] in dates:
                        cell['value'] = (datetime.strptime(cell['value'], '%Y-%m-%dT%H:%M:%SZ') + timedelta(
                            hours=1 if self.creds.get('dayLight') else 2)).strftime('%Y-%m-%dT%H:%M:%SZ')
                    fields[cell['colid']] = cell['value']
                allFields.append(fields)
        return headers, allFields

    def requestFinal(self, type_, query, param):
        self.headers['Content-Type'] = 'text/xml;charset="UTF-8"'
        data = f"""<?xml version="1.0" encoding="utf-8" standalone="yes"?><message xmlns="http://www.logobject.ch/schema-ns/jmsoverxml"><meta><type>{type_}</type><expectResponse>true</expectResponse></meta><properties><property type="string" name="username" value="{self.user}" /><property type="string" name="token" value="{self.session}" /><property type="string" name="userprofile" value="{self.profile}" /><property type="string" name="txtimeout" value="60" /></properties><body><entity class="{query}" xmlns="">{param}</entity></body></message>"""
        with post(f'https://mlog.unitymedia.de/jmsbridge/jms;{type_}?timeout=60000', headers=self.headers,
                  data=data) as r:
            return '"errorMessage"' not in r.text

    def search(self, term):
        data = f"""<additionalClause><item> AND ENTITYREFERENCECID = 10007  AND ENTITYREFERENCEID IN (select t.id from TBTASK t, TBIDCPROFILE prof, TBIDCPARTITEM idcpart where prof.userprofile_id=:profileId and prof.idcpartition_id=idcpart.partition_id and idcpart.businessunit_id=t.businessunit_id  and idcpart.region_id=t.region_id) </item></additionalClause><additionalOrder><item> order by ID DESC NULLS LAST</item></additionalOrder><parameters><param name="msgContent" type="string">{term}</param></parameters>"""
        return self.request("ch.logobject.dart.mlogistics.supervisionManagement.QrDeadMessages", data)

    def getHid(self, ID):
        data = f"""<additionalClause><item> and TASKID IN (:taskIds_EXT)</item></additionalClause><parameters><param name="taskIds_EXT" type="intArray">{ID}</param></parameters>"""
        headers, fields = self.request("ch.logobject.dart.mlogistics.taskManagement.QrOrderFullSearchWithMissing", data)
        if fields:
            return fields[0][headers['TASKHID']], fields[0][headers['STATEVALUE']]

    def refresh(self, ID):
        data = f"""<field name="taskId" value="{ID}" />"""
        query = "ch.logobject.um.bl.msg.PollActivitiesRequest"
        type_ = "wispPollActivities"
        self.requestFinal(type_, query, data)

    def checkStatus(self, HID, ID, code, status):
        data = f""" <parameters><param name="taskHid" type="string">{HID}</param><param name="taskId" type="int">{ID}</param></parameters>"""
        headers, fields = self.request("ch.logobject.dart.unitymedia.taskManagement.QrUMOrderInfoWorkLog", data)
        statusOK = False
        for field in fields:
            if field[headers['ACTIVITYCODE']] == code and field[headers['ACTIVITYSTATUS']] == status:
                statusOK = True
                break
        return statusOK

    def recover(self, MSGID):
        data = f"""<field name="messageId" value="{MSGID}" />"""
        type_ = "recoverDeadMessage"
        query = "ch.logobject.common.om.logging.DeadMessage"
        return self.requestFinal(type_, query, data)

    def discard(self, MSGID):
        data = f"""<field name="messageids" value="{{'{MSGID}'}}" />"""
        type_ = "discardDeadMessage"
        query = "ch.logobject.common.bl.msg.DiscardMessageRequest"
        return self.requestFinal(type_, query, data)

    def confirm_date(self):
        term = "CONFIRMED_DATE"
        headers, fields = self.search(term)
        for field in fields:
            ID = field[headers['ID']]
            MSGID = field[headers['MESSAGEID']]
            ERRORMSG = field[headers['ERRORMSG']]
            CREATEUSER = field[headers['CREATEUSER']]
            ENTITYINFO = field[headers['ENTITYINFO']]
            HID, _ = self.getHid(ID)
            if not HID:
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "No record Found"])
                print("[-] No record Found for:", ID)
                continue
            self.refresh(ID)
            if self.checkStatus(HID, ID, "TBS", "ACTV"):
                if self.recover(MSGID):
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "YES", "Recovered"])
                    print("\t[+] Recovered: ", ID)
                else:
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Failed to Recover"])
                    print("\t[-] Failed to Recover: ", ID)
            else:
                self.sheet.append(
                    [ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Code TBS and Status ACTV Not Matched"])
                print("\t[-] Code TBS and Status ACTV Not Matched for:", ID)
        if fields:
            print()
        else:
            print("[-] No Error Found")
        self.book.save(self.filename)

    def moretime(self):
        term = "WO_RETURN_MORETIME"
        headers, fields = self.search(term)
        for field in fields:
            ID = field[headers['ID']]
            MSGID = field[headers['MESSAGEID']]
            ERRORMSG = field[headers['ERRORMSG']]
            CREATEUSER = field[headers['CREATEUSER']]
            ENTITYINFO = field[headers['ENTITYINFO']]

            HID, status = self.getHid(ID)
            if not HID or int(status) != 160:
                self.sheet.append(
                    [ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "No record Found or Status not 'ON Hold'"])
                print("\t[-] No record Found or Status not 'ON Hold' for:", ID)
                continue
            self.refresh(ID)
            if self.checkStatus(HID, ID, "HPP", "ACTV"):
                if self.discard(MSGID):
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "YES", "Discarded"])
                    print("\t[+] Discarded: ", ID)
                else:
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Failed to Discard"])
                    print("\t[-] Failed to Discard: ", ID)
            else:
                self.sheet.append(
                    [ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Code HPP and Status ACTV Not Matched"])
                print("\t[-] Code HPP and Status ACTV Not Matched for:", ID)
        if fields:
            print()
        else:
            print("[-] No Error Found")
        self.book.save(self.filename)

    def returnQ(self, q):
        term = f"WO_RETURN_Q{q}"
        headers, fields = self.search(term)
        for field in fields:
            ID = field[headers['ID']]
            MSGID = field[headers['MESSAGEID']]
            ERRORMSG = field[headers['ERRORMSG']]
            CREATEUSER = field[headers['CREATEUSER']]
            ENTITYINFO = field[headers['ENTITYINFO']]

            HID, _ = self.getHid(ID)
            if not HID:
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "No record Found"])
                print("[-] No record Found for:", ID)
                continue
            self.refresh(ID)
            if self.checkStatus(HID, ID, "HPP", "ACTV"):
                if self.discard(MSGID):
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "YES", "Discarded"])
                    print("\t[+] Discarded: ", ID)
                else:
                    self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Failed to Discard"])
                    print("\t[-] Failed to Discard: ", ID)
            else:
                self.sheet.append(
                    [ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Code HPP and ACTV Status Not Matched"])
                print("\t[-] Code and Status Not Matched for:", ID)
        if fields:
            print()
        else:
            print("[-] No Error Found")
        self.book.save(self.filename)

    def recover_all(self, term):
        headers, fields = self.search(term)
        for field in fields:
            ID = field[headers['ID']]
            MSGID = field[headers['MESSAGEID']]
            ERRORMSG = field[headers['ERRORMSG']]
            CREATEUSER = field[headers['CREATEUSER']]
            ENTITYINFO = field[headers['ENTITYINFO']]
            HID, _ = self.getHid(ID)
            if not HID:
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "No record Found"])
                print("[-] No record Found for:", ID)
                continue

            if self.recover(MSGID):
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "YES", "Recovered"])
                print("\t[+] Recovered: ", ID)
            else:
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Failed to Recover"])
                print("\t[-] Failed to Recover: ", ID)
        if fields:
            print()
        else:
            print("[-] No Error Found")
        self.book.save(self.filename)

    def discard_all(self, term):
        headers, fields = self.search(term)
        for field in fields:
            ID = field[headers['ID']]
            MSGID = field[headers['MESSAGEID']]
            ERRORMSG = field[headers['ERRORMSG']]
            CREATEUSER = field[headers['CREATEUSER']]
            ENTITYINFO = field[headers['ENTITYINFO']]
            HID, _ = self.getHid(ID)
            if not HID:
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "No record Found"])
                print("[-] No record Found for:", ID)
                continue
            if self.discard(MSGID):
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "YES", "Discarded"])
                print("\t[+] Discarded: ", ID)
            else:
                print("\t[-] Failed to Discard: ", ID)
                self.sheet.append([ID, ERRORMSG, CREATEUSER, ENTITYINFO, term, "NO", "Failed to Discard"])
        if fields:
            print()
        else:
            print("[-] No Error Found")
        self.book.save(self.filename)

    def clicks(self, query, data, term):
        headers, fields = self.request(query, data)
        beforeLeng = len(fields)
        print(f"[+] Orders before: {beforeLeng}")
        for field in fields:
            ID = field[headers['TASKID']]
            self.refresh(ID)
            print("\t[+] Refreshed:" + ID)
        headers, fields = self.request(query, data)
        afterLeng = len(fields)
        self.sheet2.append([term, beforeLeng, afterLeng])
        print(f"[+] Orders After: {afterLeng}")
        self.book.save(self.filename)

    def getCredentials(self):
        self.creds = {line.split('=')[0]: line.split('=')[1] for line in open("credentials.txt").read().splitlines()}
        return self.creds


def main():
    auto = Automate()

    print("[+] Getting Credentials")
    try:
        creds = auto.getCredentials()
    except Exception as e:
        Mbox("Failed to Credentials", "credentials.txt file is missing. Please create it.", 0)
        return

    print("[+] Logging in to the server")
    try:
        token, profile = auto.login(creds)
        print("[>] Signed as:", creds.get('username'))
        print("[>] Profile Selected:", profile)
    except KeyError:
        print("[-] Login Invalid, Please Check your Login Details")
        Mbox("Failed to Login", "Login Invalid, Please Check your Login Details", 0)
        return
    except exceptions.ConnectionError:
        print("[-] Server Down, Failed to Login")
        Mbox("Failed to Login", "Server Down, Failed to Login", 0)
        return
    except exceptions.ReadTimeout:
        print("[-] Server Down, Failed to Login")
        Mbox("Failed to Login", "Server Down, Failed to Login", 0)
        return

    print("\n[>] Term: CONFIRMED_DATE")
    try:
        auto.confirm_date()
    except Exception as e:
        print("[-] Something went Wrong")

    for term in ['CONFIRM_SCH']:
        print("\n[>] Term:", term)
        try:
            auto.recover_all(term)
        except Exception as e:
            print("[-] Something went Wrong")

    for term in ['Number of comments exceeded configured limit', 'RETURN_SIG', 'RESCH_WORK_ORDER', 'TECH_ON_SITE']:
        print("\n[>] Term: ", term)
        try:
            auto.discard_all(term)
        except Exception as e:
            print("[-] Something went Wrong")

    print("\n[>] Term: WO_RETURN_MORETIME")
    try:
        auto.moretime()
    except Exception as e:
        print("[-] Something went Wrong")

    for q in [5, 6]:
        print(f"\n[>] Term: WO_RETURN_Q{q}")
        try:
            auto.returnQ(q)
        except Exception as e:
            print("[-] Something went Wrong")

    for click in [["Term Ungleich TVO",
                   """<additionalClause><item> and STATEVALUE IN (:SP_STATEVALUE)</item><item> and UPPER(CURRENTACTIVITY) &lt;> UPPER(:SP_CURRENTACTIVITY1)</item><item> and qRes.SCHEDGROUPTYPE IN (:schedGroupType) and qRes.STATEVALUE IN (:taskStateValue)</item></additionalClause><parameters><param name="SP_STATEVALUE" type="stringArray">80,220,60</param><param name="SP_CURRENTACTIVITY1" type="string">TVO</param><param name="schedGroupType" type="intArray">0,20,60,80</param><param name="taskStateValue" type="intArray">30,40,50,60,70,80,85,90,100,110,120,130,140,145,150,160,166,170,180,190,220,225,230,240,175,185,250</param></parameters>"""],
                  ["Zu Term ungleich TBS",
                   """<additionalClause><item> and STATEVALUE IN (:SP_STATEVALUE)</item><item> and (UPPER(CURRENTACTIVITY) NOT LIKE :SP_CURRENTACTIVITY0 and UPPER(CURRENTACTIVITY) NOT LIKE :SP_CURRENTACTIVITY1 and UPPER(CURRENTACTIVITY) NOT LIKE :SP_CURRENTACTIVITY2 and UPPER(CURRENTACTIVITY) NOT LIKE :SP_CURRENTACTIVITY3 and UPPER(CURRENTACTIVITY) NOT LIKE :SP_CURRENTACTIVITY4 OR UPPER(CURRENTACTIVITY) IS NULL )</item><item> and qRes.SCHEDGROUPTYPE IN (:schedGroupType) and qRes.STATEVALUE IN (:taskStateValue)</item></additionalClause><parameters><param name="SP_STATEVALUE" type="stringArray">50,30,160</param><param name="SP_CURRENTACTIVITY0" type="string">%TBS%</param><param name="SP_CURRENTACTIVITY1" type="string">%HPP%</param><param name="SP_CURRENTACTIVITY2" type="string">%*SMD%</param><param name="SP_CURRENTACTIVITY3" type="string">%MTW%</param><param name="SP_CURRENTACTIVITY4" type="string">%*PUO%</param><param name="schedGroupType" type="intArray">0,20,60,80</param><param name="taskStateValue" type="intArray">30,40,50,60,70,80,85,90,100,110,120,130,140,145,150,160,166,170,180,190,220,225,230,240,175,185,250</param></parameters>"""]]:
        print("\n[>] Clicking:", click[0])
        query = "ch.logobject.dart.unitymedia.taskManagement.QrUMOrderOverview"
        data = click[1]
        try:
            auto.clicks(query, data, click[0])
        except Exception as e:
            print("[-] Something went Wrong")
    print("[+] Completed")
    auto.book.save(auto.filename)
    time.sleep(10)


def Mbox(title, text, style):
    return ctypes.windll.user32.MessageBoxW(0, text, title, style)


if __name__ == '__main__':
    try:
        main()
    except Exception as e:
        Mbox("Failed to Run the script", f"Please report the following incident\n{e}", 0)
